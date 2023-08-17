from django.shortcuts import render, HttpResponse, redirect
from django.contrib import messages
from django.core.mail import send_mail, EmailMultiAlternatives
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from .models import Event, EventPass, EventSlot
from ..users.models import User
import random
import string
import json
from datetime import date, time, datetime
from decouple import config

from ..users.utils import account_activation_token
import random, string
import openpyxl
from .models import User
from ..users.manager import generate_qr, generate_user_secure_id
from ..users.views import *
from datetime import datetime
import json

from ..hoods.models import Hood
from django.db.utils import IntegrityError

global counters
counters = {} 
for event in Event.objects.all():
    counters[event.name]=event.passes_generated
for slot in EventSlot.objects.all():
    counters[slot.slot_id] = slot.passes_generated


@login_required
def events_home(request):
    events = Event.objects.filter(is_display=True)
    past_events = []
    events_today = []
    for event in events:
        if event.date < date.today():
            past_events.append(event)
            events = events.exclude(event_id=event.event_id)
        elif event.date == date.today():
            events_today.append(event)
            events = events.exclude(event_id=event.event_id)
    past_events = sorted(past_events, key=lambda x: [x.date, convert_time_to_start_time(x.time)])[::-1]
    events_sorted = sorted(events, key=lambda x: [x.date, convert_time_to_start_time(x.time)])
    events_today = sorted(events_today, key=lambda x: [x.date, convert_time_to_start_time(x.time)])
    if len(events_today) > 0:
        live_event = events_today.pop(0)
        for event in events_today:
            events_sorted.append(event)
    else:
        live_event = None

    events_sorted = sorted(events, key=lambda x: [x.date, convert_time_to_start_time(x.time)])

    event_slots = EventSlot.objects.all()
    user_passes = EventPass.objects.filter(user_id=request.user)
    user_passes = sorted(user_passes, key=lambda x: [x.event_id.date])[::-1]

    return render(request, 'events.html', {'upcoming_events':events_sorted, 'past_events':past_events, 'events_today':events_today, 'event_slots':event_slots, 'user_passes':user_passes, 'live_event':live_event})			  


@csrf_exempt
@login_required
def generate_pass(request, event_id, slot_id=None):
    user = User.objects.get(registration_id=request.user.registration_id)
    event = Event.objects.get(event_id=event_id)
    if event.is_display == False:
        data = {
                'status':False,
                'message':'Event is not available'
                }
        return HttpResponse(json.dumps(data))
    if not event.is_booking or not event.booking_required:
        data = {
                'status':False,
                'message':'Booking is currently not available for this event'
                }
        return HttpResponse(json.dumps(data))
    if event.booking_complete:
        data = {
                'status':False,
                'message':'Sold out! Please try again later.'
                }
        return HttpResponse(json.dumps(data))
    
    
    ### FOR LINKING TWO EVENTS ###

    # if event.event_id == 'Whodunit [Main Audi]@Frosh23' or event.event_id == 'Whodunit [Tan Audi]@Frosh23':
    #     orientation1 = Event.objects.get(event_id='Whodunit [Main Audi]@Frosh23')
    #     orientation2 = Event.objects.get(event_id='Whodunit [Tan Audi]@Frosh23')
    #     if not EventPass.objects.filter(user_id=user, event_id=orientation1).count() and not EventPass.objects.filter(user_id=user, event_id=orientation2).count():
    #         if slot_id and event.slot_id==slot_id:
    #             event_slot = EventSlot.objects.get(slot_id=slot_id)
    #             counters[slot_id] += 1
    #             event_slot.refresh_from_db()
    #             if counters[slot_id]<=event_slot.max_capacity:
    #                 while True:
    #                     pass_id = ''.join(random.choices(string.ascii_uppercase +
    #                                         string.digits, k=16))

    #                     if not EventPass.objects.filter(pass_id=pass_id).count():
    #                         break
    #                 generated_pass = EventPass(event_id=event_slot.event,slot_id=event_slot ,pass_id=pass_id, user_id=user)
    #                 generated_pass.qr = user.qr
    #                 generated_pass.time = event_slot.time
    #                 generated_pass.save()
    #                 # confirmation_email(generated_pass=generated_pass)

    #                 event_slot.passes_generated = counters[slot_id]
    #                 user.events.append(str(event))
    #                 user.save()
    #                 event_slot.save()
    #                 data = {
    #                 'pass_qr':generated_pass.qr,
    #                 'status':True
    #                 }
    #                 return HttpResponse(json.dumps(data))
    #             else:
    #                 data={
    #                     'status':False,
    #                     'message':"There was an error!"
    #                 }
    #                 counters[slot_id]-=1
    #                 return HttpResponse(json.dumps(data))
    #         elif slot_id and event.slot_id!=slot_id:
    #             data={
    #                     'status':False,
    #                     'message':"There was an error!"
    #                 }
    #             return HttpResponse(json.dumps(data))
    #         else:
    #             counters[event.name]+=1
    #             event.refresh_from_db()
    #             if counters[event.name]<=event.max_capacity:
    #                 while True:
    #                     pass_id = ''.join(random.choices(string.ascii_uppercase +
    #                                         string.digits, k=16))

    #                     if not EventPass.objects.filter(pass_id=pass_id).count():
    #                         break

    #                 generated_pass = EventPass(event_id=event, pass_id=pass_id, user_id=user)
    #                 generated_pass.qr = user.qr
    #                 generated_pass.time = event.time
    #                 generated_pass.save()
                    
    #                 # confirmation_email(generated_pass=generated_pass)
    #                 event.passes_generated = counters[event.name]
    #                 user.events.append(str(event))
    #                 user.save()
    #                 event.save()
    #                 data = {
    #                     'pass_qr':generated_pass.qr,
    #                     'status':True
    #                 }
    #                 return HttpResponse(json.dumps(data))
    #             else:
    #                 data={
    #                     'status':False,
    #                     'message':"There was an error!"
    #                 }
    #                 counters[event.name]-=1
    #                 return HttpResponse(json.dumps(data))
    #     else:
    #         data={
    #                 'status':False,
    #                 'message':"You have already booked a pass for this event!"
    #             }
    #         return HttpResponse(json.dumps(data))

    if not EventPass.objects.filter(user_id=user, event_id=event).count():
        if slot_id and event.slot_id==slot_id:
            event_slot = EventSlot.objects.get(slot_id=slot_id)
            counters[slot_id] += 1
            event_slot.refresh_from_db()
            if counters[slot_id]<=event_slot.max_capacity:
                while True:
                    pass_id = ''.join(random.choices(string.ascii_uppercase +
                                        string.digits, k=16))

                    if not EventPass.objects.filter(pass_id=pass_id).count():
                        break
                generated_pass = EventPass(event_id=event_slot.event,slot_id=event_slot ,pass_id=pass_id, user_id=user)
                generated_pass.qr = user.qr
                generated_pass.time = event_slot.time
                generated_pass.save()
                # confirmation_email(generated_pass=generated_pass)

                event_slot.passes_generated = counters[slot_id]
                user.events.append(str(event))
                user.save()
                event_slot.save()
                data = {
                'pass_qr':generated_pass.qr,
                'status':True
                }
                return HttpResponse(json.dumps(data))
            else:
                data={
                    'status':False,
                    'message':"No more passes available for this slot!"
                }
                counters[slot_id]-=1
                # event.booking_complete = True
                # event.is_booking = False
                # event.save()
                return HttpResponse(json.dumps(data))
        elif slot_id and event.slot_id!=slot_id:
            data={
                    'status':False,
                    'message':"There was an error!"
                }
            return HttpResponse(json.dumps(data))
        else:
            counters[event.name]+=1
            event.refresh_from_db()
            if counters[event.name]<=event.max_capacity:
                while True:
                    pass_id = ''.join(random.choices(string.ascii_uppercase +
                                        string.digits, k=16))

                    if not EventPass.objects.filter(pass_id=pass_id).count():
                        break

                generated_pass = EventPass(event_id=event, pass_id=pass_id, user_id=user)
                generated_pass.qr = user.qr
                generated_pass.time = event.time
                generated_pass.save()
                
                # confirmation_email(generated_pass=generated_pass)
                event.passes_generated = counters[event.name]
                user.events.append(str(event))
                user.save()
                event.save()
                data = {
                    'pass_qr':generated_pass.qr,
                    'status':True
                }
                return HttpResponse(json.dumps(data))
            else:
                data={
                    'status':False,
                    'message':"No more passes available for this event!"
                }
                # event.booking_complete = True
                # event.is_booking = False
                # event.save()
                counters[event.name]-=1
                return HttpResponse(json.dumps(data))
    else:
        data={
                    'status':False,
                    'message':"You have already booked a pass for this event!"
                }
        return HttpResponse(json.dumps(data))


def confirmation_email(generated_pass:EventPass):
    subject= subject = f'Registration successful for {generated_pass.event_id} | Frosh 23'
    from_email = settings.EMAIL_HOST_USER
    to = generated_pass.user_id.email
    text_content = f'Hi {generated_pass.user_id}, thank you for registering in {generated_pass.event_id}. Time is {generated_pass.time}'
    html_content = render_to_string('email_event.html', { 'event_name':generated_pass.event_id.name,'venue':generated_pass.event_id.venue, 'user_name':f'{generated_pass.user_id.first_name}', 'pass_id':generated_pass.pass_id, 'pass_time':generated_pass.time, 'qr':generated_pass.qr })
    msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
    msg.attach_alternative(html_content, "text/html")
    msg.send(fail_silently=False)


def convert_time_to_start_time(time_string):
    split = time_string.split('-')
    start_time = datetime.strptime(split[0].strip(), "%H:%M:%S")
    # start_date = datetime.strptime(date, )
    # datetime.combine(start_date, start_time)
    end_time =  datetime.strptime(''.join(char for char in split[1] if char!=' '), "%H:%M:%S") 
    return start_time


def user_count():
    active_users = User.objects.filter(is_active=True)
    staff_users = User.objects.filter(is_staff=True)
    superusers = User.objects.filter(is_superuser=True)
    print('Active Users: ',active_users.count())
    print('Staff Users: ',staff_users.count())
    print('Superusers: ',superusers.count())


def tslas_checkup():
    all_users = EventPass.objects.filter(event_id=Event.objects.get(event_id='Whodunit [Main Audi]@Frosh23'))
    for user in all_users:
        if user.user_id.email.find('blas23@thapar.edu') != -1:
            print(f"[Main Auditorium] {user.user_id.registration_id} {user.user_id} {user.user_id.email}")
    all_users = EventPass.objects.filter(event_id=Event.objects.get(event_id='Whodunit [Tan Audi]@Frosh23'))
    for user in all_users:
        if user.user_id.email.find('blas23@thapar.edu') != -1:
            print(f"[Tan Auditorium] {user.user_id.registration_id} {user.user_id} {user.user_id.email}")


def check_count():
    while True:
        print(EventPass.objects.filter(event_id=Event.objects.get(event_id='Event X@Frosh23')).count())


def generate_password():
    user = User.objects.get(registration_id='141208')
    user.set_password('12345678')
    print(user)


def gen_special_passed():
    count = 0
    wb_obj = openpyxl.load_workbook()
    sheet = wb_obj.active
    start = int(input("Start: "))
    end = int(input('End: '))
    event = Event.objects.get(event_id='Event X@Frosh23')
    print(event)
    for row in range(start, end+1):
        data = []
        user = User()
        for column in range(1,7):
            cell = sheet.cell(row=row, column=column)
            data.append(cell.value)
        user = User.objects.create_user(registration_id=int(data[0]),first_name = data[1], last_name = data[3] if data[3] else ' ',
        email = data[4], is_staff = False, is_superuser = False, is_active = False, image=data[5])
        count+=1
        user.secure_id = generate_user_secure_id()
        user.qr = upload_qr(generate_qr(json.dumps({
                'registration_id':user.registration_id,
                'secure_id':user.secure_id
            }), user.registration_id), user.secure_id)
        print(f"[{count}] ",f"[ADDED TO DATABASE] {user.first_name}")
        while True:
            pass_id = ''.join(random.choices(string.ascii_uppercase +
                                string.digits, k=16))

            if not EventPass.objects.filter(pass_id=pass_id).count():
                break

        generated_pass = EventPass(event_id=event, pass_id=pass_id, user_id=user)
        generated_pass.qr = user.qr
        generated_pass.time = event.time
        generated_pass.save()
        
        confirmation_email(generated_pass=generated_pass)
        print(pass_id, 'emailed')
        event.passes_generated = counters[event.name]
        user.events.append(str(event))
        user.save()
        event.save()





def database_migrate():
    # users = User.objects.all()
    # print(users)
    # events = Event.objects.all()
    # print(events)
    # hoods = Hood.objects.all()
    # print(hoods)
    # event_slots = EventSlot.objects.all()
    # print(event_slots)
    # passes = EventPass.objects.all()
    # print(passes)


    # wb = openpyxl.Workbook()
    # sheet = wb.active
    # count = 0
    # for user in users:
    #     data = [user.registration_id, user.first_name, user.last_name, user.email, user.image, user.secure_id, user.qr, str(user.events), str(user.hood), user.is_active, user.is_staff, user.is_superuser, str(user.password)]
    #     print(count, user)
    #     count+=1
    #     sheet.append(data)

    # for event in events:
    #     data = [event.event_id,event.name, event.image,event.description,event.venue, event.date, event.time, event.max_capacity, event.passes_generated, event.calendar_url, str(event.slot_id), event.is_booking, event.is_display, event.slots_required, event.booking_complete, event.booking_required]
    #     print(count, event)
    #     count+=1
    #     sheet.append(data)

    # for event in event_slots:
    #     data = [event.slot_id,str(event.event),event.venue, event.date, event.time, event.max_capacity, event.passes_generated, event.calendar_url,  event.booking_complete]
    #     print(count, event)
    #     count+=1
    #     sheet.append(data)


    # for hood in hoods:
    #     data = [hood.id, hood.name,hood.description, hood.image, hood.member_count ]
    #     print(count, hood)
    #     count+=1
    #     sheet.append(data)

    # for passs in passes:
    #     data = [passs.pass_id, str(passs.event_id), str(passs.slot_id), str(passs.user_id.registration_id), passs.qr, passs.entry_status, passs.time]
    #     print(passs)
    #     sheet.append(data)

    # writing values to cells
    # c1.value = "ANKIT"

    # c2 = sheet.cell(row= 1 , column = 2)
    # c2.value = "RAI"

    # # Once have a Worksheet object, one can
    # # access a cell object by its name also.
    # # A2 means column = 1 & row = 2.
    # c3 = sheet['A2']
    # c3.value = "RAHUL"

    # # B2 means column = 2 & row = 2.
    # c4 = sheet['B2']
    # c4.value = "RAI"

    # # Anytime you modify the Workbook object
    # # or its sheets and cells, the spreadsheet
    # # file will not be saved until you call
    # # the save() workbook method.
    # wb.save(r"C:\Users\Pancham\Desktop\projects\frosh-web-production\database_users.xlsx")
    # wb.save(r"C:\Users\Pancham\Desktop\projects\frosh-web-production\database_events.xlsx")
    # wb.save(r"C:\Users\Pancham\Desktop\projects\frosh-web-production\database_events_slots.xlsx")
    # wb.save(r"C:\Users\Pancham\Desktop\projects\frosh-web-production\database_hoods.xlsx")
    # wb.save(r"C:\Users\Pancham\Desktop\projects\frosh-web-production\database_passes.xlsx")

    wb = openpyxl.load_workbook(r"C:\Users\Pancham\Desktop\projects\frosh-web-production\database_users.xlsx")
    sheet = wb.active
    count = 0
    for row in range(19716, sheet.max_row+1):
        data = []
        for column in range(1,14):
            cell = sheet.cell(row=row, column=column)
            data.append(cell.value)
        try:
#     data = [user.registration_id, user.first_name, user.last_name, user.email, user.image, user.secure_id, user.qr, str(user.events), str(user.hood), user.is_active, user.is_staff, user.is_superuser, str(user.password)]
            user = User(registration_id=int(data[0]),first_name = data[1], last_name = data[2], email = data[3], is_staff = data[10], is_superuser = data[11], is_active = data[9], image=data[4])
            user.secure_id = data[5]
            user.qr = data[6] if data[6] else ' '
            user.events = list(data[7]) if data[7] else []
            hood = Hood.objects.filter(name=data[8])
            user.hood = hood[0] if hood else None
            user.password = data[12]
            user.save()
            count+=1
            print(count, user.registration_id)
        except IntegrityError:
            pass
    
    # load the events workbook and sheet to add all events to database
    # wb = openpyxl.load_workbook(r"C:\Users\Pancham\Desktop\projects\frosh-web-production\database_events.xlsx")
    # sheet = wb.active
    # for row in range(1, 2):
    #     data = []
    #     for column in range(1,18):
    #         cell = sheet.cell(row=row, column=column)
    #         data.append(cell.value)
    #     #data = [event.event_id,event.name, event.image,event.description,event.venue, event.date, event.time, event.max_capacity, event.passes_generated, event.calendar_url, str(event.slot_id), event.is_booking, event.is_display, event.slots_required, event.booking_complete, event.booking_required]
    #     event = Event(event_id=data[0], name=data[1], image=data[2], description=data[3], venue=data[4], date=data[5], time=data[6], max_capacity=data[7], passes_generated=data[8], calendar_url=data[9], is_booking=data[11], is_display=data[12], slots_required=data[13], booking_complete=data[14], booking_required=data[15])
    #     event.save()
    #     print(event)

    #load the hoods workbook and sheet to add all hoods to database
    # wb = openpyxl.load_workbook(r"C:\Users\Pancham\Desktop\projects\frosh-web-production\database_hoods.xlsx")
    # sheet = wb.active
    # for row in range(1, sheet.max_row+1):
    #     data = []
    #     for column in range(1,6):
    #         cell = sheet.cell(row=row, column=column)
    #         data.append(cell.value)
    #     #data = [hood.id, hood.name,hood.description, hood.image, hood.member_count ]
    #     hood = Hood(id=data[0], name=data[1], description=data[2] if data[2] else ' ', image=data[3], member_count=data[4])
    #     hood.save()
    #     print(hood)


        
# database_migrate()
    
# def ticket_count():
#     while True:
#         slot1 = EventPass.objects.filter(event_id=Event.objects.get(event_id='Elyserra@Frosh23')).exclude(slot_id=EventSlot.objects.get(slot_id='D35OXQJ8WY48JTAF'))
#         slot2 = EventPass.objects.filter(slot_id=EventSlot.objects.get(slot_id='D35OXQJ8WY48JTAF'))

#         print({
#             'Slot 1': slot1.count(),
#             'Slot 2': slot2.count()
#         })

# ticket_count()